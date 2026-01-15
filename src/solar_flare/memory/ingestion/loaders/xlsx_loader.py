"""
Excel document loader using openpyxl.
"""
from pathlib import Path
from typing import BinaryIO, Dict, Any

from .base import BaseDocumentLoader, LoadedDocument, DocumentLoadError


class XlsxLoader(BaseDocumentLoader):
    """Load Excel documents using openpyxl."""

    SUPPORTED_EXTENSIONS = [".xlsx"]

    def __init__(self, max_rows: int = 1000, include_formulas: bool = False):
        """
        Initialize Excel loader.

        Args:
            max_rows: Maximum rows to extract per sheet
            include_formulas: Whether to include formula text (vs. values only)
        """
        self.max_rows = max_rows
        self.include_formulas = include_formulas

    def load(self, file_path: Path) -> LoadedDocument:
        """Load Excel from file path."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise DocumentLoadError(
                "openpyxl not installed. Run: pip install openpyxl",
                str(file_path)
            )

        try:
            wb = load_workbook(file_path, data_only=not self.include_formulas)
            return self._extract_content(wb, str(file_path), file_path.name)
        except Exception as e:
            raise DocumentLoadError(f"Failed to load Excel: {e}", str(file_path), e)

    def load_from_stream(
        self,
        stream: BinaryIO,
        filename: str,
        source_path: str,
    ) -> LoadedDocument:
        """Load Excel from binary stream."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise DocumentLoadError(
                "openpyxl not installed. Run: pip install openpyxl",
                source_path
            )

        try:
            wb = load_workbook(stream, data_only=not self.include_formulas)
            return self._extract_content(wb, source_path, filename)
        except Exception as e:
            raise DocumentLoadError(f"Failed to load Excel stream: {e}", source_path, e)

    def _extract_content(self, wb, source_path: str, filename: str) -> LoadedDocument:
        """Extract content from openpyxl Workbook."""
        text_parts = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_text = [f"--- Sheet: {sheet_name} ---"]

            row_count = 0
            for row in sheet.iter_rows(values_only=True):
                if row_count >= self.max_rows:
                    sheet_text.append(f"[... truncated at {self.max_rows} rows ...]")
                    break

                # Convert row to text, handling None values
                row_values = []
                has_content = False
                for cell in row:
                    if cell is not None:
                        row_values.append(str(cell))
                        has_content = True
                    else:
                        row_values.append("")

                if has_content:
                    row_text = " | ".join(row_values)
                    sheet_text.append(row_text)
                    row_count += 1

            text_parts.append("\n".join(sheet_text))

        content = "\n\n".join(text_parts)

        metadata: Dict[str, Any] = {
            "filename": filename,
            "sheet_names": wb.sheetnames,
            "sheet_count": len(wb.sheetnames),
        }

        title = self._generate_title(filename)

        return LoadedDocument(
            content=content,
            title=title,
            source_path=source_path,
            file_format="xlsx",
            content_hash=LoadedDocument.compute_hash(content),
            metadata=metadata,
        )
